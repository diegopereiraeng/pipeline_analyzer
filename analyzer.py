import requests
import json
import yaml
from collections import defaultdict
import csv
import time
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
import pandas as pd
import logging
from dotenv import load_dotenv, find_dotenv
import os

_ = load_dotenv(override=True)

# Setup logging
logging.basicConfig(filename='pipeline_analysis.log', level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

# Replace with your actual API key
API_KEY = os.getenv('API_KEY')
HARNESS_ACCOUNT_ID = os.getenv('HARNESS_ACCOUNT_ID')
BASE_URL = 'https://app.harness.io/gateway'
DEBUG = False
DEBUG_PIPELINE_NAME = "Post_PR_Release_Branch"

headers = {
    'Authorization': f'Bearer {API_KEY}',
    'Accept': 'application/json',
    'Content-Type': 'application/json'
}

# Timer function
def timer_func(func):
    def wrapper(*args, **kwargs):
        start_time = time.time()
        result = func(*args, **kwargs)
        end_time = time.time()
        print(f"{func.__name__} took {end_time - start_time:.2f} seconds")
        return result
    return wrapper

@timer_func
def get_orgs():
    url = f'{BASE_URL}/ng/api/organizations?accountIdentifier={HARNESS_ACCOUNT_ID}&pageSize=500'
    print(f'Fetching orgs: {url}')
    response = requests.get(url, headers=headers)
    response.raise_for_status()
    return response.json()['data']['content']

@timer_func
def get_projects(org_identifier):
    url = f'{BASE_URL}/ng/api/aggregate/projects?routingId={HARNESS_ACCOUNT_ID}&accountIdentifier={HARNESS_ACCOUNT_ID}&orgIdentifier={org_identifier}&pageIndex=0&pageSize=500&sortOrders=createdAt%2CDESC'
    print(f'Fetching projects for org {org_identifier}: {url}')
    response = requests.get(url, headers=headers)
    response.raise_for_status()
    return response.json()['data']['content']

@timer_func
def get_pipelines(org_identifier, project_identifier):
    url = f'{BASE_URL}/pipeline/api/pipelines/list?routingId={HARNESS_ACCOUNT_ID}&accountIdentifier={HARNESS_ACCOUNT_ID}&projectIdentifier={project_identifier}&orgIdentifier={org_identifier}&page=0&sort=lastUpdatedAt%2CDESC&size=500'
    data = {
        "filterType": "PipelineSetup"
    }
    response = requests.post(url, headers=headers, json=data)  # Fix this line
    try:
        response.raise_for_status()
    except requests.exceptions.HTTPError as e:
        print(f'Error fetching pipelines: {e}\nURL: {url}\nResponse: {response.text}')
        return None
    return response.json()['data']['content']

@timer_func
def get_pipeline_yaml(org_identifier, project_identifier, pipeline_identifier, store_type, connector_ref=None, repo_name=None):
    if store_type == "INLINE":
        url = f'{BASE_URL}/pipeline/api/pipelines/{pipeline_identifier}?accountIdentifier={HARNESS_ACCOUNT_ID}&orgIdentifier={org_identifier}&projectIdentifier={project_identifier}&validateAsync=true'
    else:
        url = f'{BASE_URL}/pipeline/api/pipelines/{pipeline_identifier}?accountIdentifier={HARNESS_ACCOUNT_ID}&orgIdentifier={org_identifier}&projectIdentifier={project_identifier}&validateAsync=true&loadFromFallbackBranch=true&parentEntityConnectorRef={connector_ref}&parentEntityRepoName={repo_name}'
    
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        yaml_pipeline = response.json()['data']['yamlPipeline']
        try:
            parsed_yaml = yaml.safe_load(yaml_pipeline)
        except yaml.YAMLError as yaml_error:
            logging.error(f'Error parsing YAML: {yaml_error}\nYAML content: {yaml_pipeline}')
            return None, f'Error parsing YAML: {yaml_error}'
        return parsed_yaml, None
    except requests.exceptions.HTTPError as e:
        logging.error(f'Error fetching pipeline YAML: {e}\nURL: {url}\nResponse: {response.text}')
        return None, str(e)
    except requests.exceptions.ConnectionError as e:
        logging.error(f'Connection error: {e}\nURL: {url}')
        return None, str(e)


@timer_func
def get_template_yaml(template_ref, version_label='0.0.1', current_level='account', org_identifier=None, project_identifier=None, parent_pipeline_id=None):
    if template_ref.startswith('account.'):
        template_id = template_ref.replace('account.', '')
        url = f'{BASE_URL}/template/api/templates/{template_id}?accountIdentifier={HARNESS_ACCOUNT_ID}&versionLabel={version_label}&loadFromFallbackBranch=true'
    elif template_ref.startswith('org.'):
        template_id = template_ref.replace('org.', '')
        url = f'{BASE_URL}/template/api/templates/{template_id}?accountIdentifier={HARNESS_ACCOUNT_ID}&orgIdentifier={org_identifier}&loadFromFallbackBranch=true'
    else:
        # Default to current level logic
        if current_level == 'project':
            url = f'{BASE_URL}/template/api/templates/{template_ref}?accountIdentifier={HARNESS_ACCOUNT_ID}&orgIdentifier={org_identifier}&projectIdentifier={project_identifier}&loadFromFallbackBranch=true'
        elif current_level == 'org':
            url = f'{BASE_URL}/template/api/templates/{template_ref}?accountIdentifier={HARNESS_ACCOUNT_ID}&orgIdentifier={org_identifier}&loadFromFallbackBranch=true'
        elif current_level == 'account':
            url = f'{BASE_URL}/template/api/templates/{template_ref}?accountIdentifier={HARNESS_ACCOUNT_ID}&versionLabel={version_label}&loadFromFallbackBranch=true'
        else:
            return None, f'Unknown level: {current_level} for template reference: {template_ref}'

    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        if parent_pipeline_id == DEBUG_PIPELINE_NAME and DEBUG == True:
            template_yaml = yaml.safe_load(response.json()['data']['yaml'])
            logging.info(f'Template YAML fetched for {template_ref}: {template_yaml}')
        return yaml.safe_load(response.json()['data']['yaml']), None
    except requests.exceptions.HTTPError as e:
        logging.error(f'Error fetching template YAML: {e}\nURL: {url}\nResponse: {response.text}')
        return None, str(e)
    except requests.exceptions.ConnectionError as e:
        logging.error(f'Connection error: {e}\nURL: {url}')
        return None, str(e)

def process_stages(stages, processed_templates, template_count, current_level='project', org_identifier=None, project_identifier=None, parent_pipeline_id=None):
    infra_types = set()
    ci_stage_count = 0
    has_template = False
    templates_used = set()

    for stage in stages:
        if 'stage' in stage:
            stage_data = stage['stage']
            if parent_pipeline_id == DEBUG_PIPELINE_NAME and DEBUG == True:
                logging.info(f'Processing stage: {stage_data["name"]}')
                logging.info(f'Stage data: {json.dumps(stage_data, indent=2)}')
            if 'template' in stage_data and 'templateRef' in stage_data['template']:
                template_ref = stage_data['template']['templateRef']
                version_label = stage_data['template'].get('versionLabel', '0.0.1')
                templates_used.add(template_ref)
                if template_ref not in processed_templates:
                    processed_templates[template_ref] = {'count': 0, 'type': None, 'ci': False, 'infra': set()}
                    template_yaml, error = get_template_yaml(template_ref, current_level=current_level, org_identifier=org_identifier, project_identifier=project_identifier, version_label=version_label, parent_pipeline_id=parent_pipeline_id)
                    if template_yaml:
                        processed_templates[template_ref]['count'] += 1
                        logging.info(f'Incremented template count for {template_ref}: {processed_templates[template_ref]["count"]}')
                        template_level = current_level
                        if template_ref.startswith('account.'):
                            template_level = 'account'
                        elif template_ref.startswith('org.'):
                            template_level = 'org'
                        if parent_pipeline_id == DEBUG_PIPELINE_NAME and DEBUG == True:
                            logging.info(f'Fetched template YAML for {template_ref}: {json.dumps(template_yaml, indent=2)}')
                        infra_types_pipeline, ci_stages_count, has_template_stage, templates_used_recursive = process_stages(
                            template_yaml.get('template', {}).get('spec', {}).get('stages', []), processed_templates, template_count, template_level, org_identifier, project_identifier, parent_pipeline_id
                        )
                        infra_types.update(infra_types_pipeline)
                        processed_templates[template_ref]['infra'].update(infra_types_pipeline)
                        ci_stage_count += ci_stages_count
                        has_template = has_template or has_template_stage
                        templates_used.update(templates_used_recursive)
                    if template_yaml and template_yaml.get('template', {}).get('type') == 'Stage':
                        if template_yaml.get('template', {}).get('spec', {}).get('type') == 'CI':
                            logging.info(f'Identified CI stage from template: {stage_data["name"]}')
                            processed_templates[template_ref]['ci'] = True
                            ci_stage_count += 1
                            infrastructure = template_yaml.get('template', {}).get('spec', {}).get('infrastructure', {'type': 'Harness Cloud'})
                            infra_type = infrastructure.get('type', 'Harness Cloud')
                            infra_types.add(infra_type)
                            processed_templates[template_ref]['infra'].add(infra_type)
                else:
                    if processed_templates[template_ref]['ci']:
                        ci_stage_count += 1
                        infra_types.update(processed_templates[template_ref]['infra'])
            if stage_data.get('type') == 'CI' or stage_data.get('templateInputs', {}).get('type') == 'CI':
                logging.info(f'Identified CI stage: {stage_data["name"]}')
                ci_stage_count += 1
                if 'spec' in stage_data:
                    infrastructure = stage_data['spec'].get('infrastructure', {'type': 'Harness Cloud'})
                    infra_type = infrastructure.get('type', 'Harness Cloud')
                    infra_types.add(infra_type)
        elif 'parallel' in stage:
            if parent_pipeline_id == DEBUG_PIPELINE_NAME and DEBUG == True:
                logging.info('Processing parallel stages')
            parallel_infra_types, parallel_ci_stage_count, parallel_has_template, templates_used_parallel = process_stages(
                stage['parallel'], processed_templates, template_count, current_level, org_identifier, project_identifier, parent_pipeline_id
            )
            infra_types.update(parallel_infra_types)
            ci_stage_count += parallel_ci_stage_count
            has_template = has_template or parallel_has_template
            templates_used.update(templates_used_parallel)

    if len(infra_types) > 1:
        infra_types = {'Mixed'}

    return infra_types, ci_stage_count, has_template, templates_used

def analyze_pipelines(pipelines, org_identifier, project_identifier, processed_templates, template_count):
    ci_stage_count = 0
    total_ci_stages = 0
    total_pipelines_with_ci = 0
    infra_types = defaultdict(int)
    pipeline_details = []
    pipeline_errors = []
    total_pipelines = len(pipelines)

    for pipeline in pipelines:
        store_type = pipeline.get('storeType', 'INLINE')
        connector_ref = pipeline.get('connectorRef')
        repo_name = pipeline.get('repoName')

        pipeline_identifier = pipeline['identifier']
        if pipeline_identifier != DEBUG_PIPELINE_NAME and DEBUG == True:
            continue
                
        pipeline_yaml, error = get_pipeline_yaml(org_identifier, project_identifier, pipeline_identifier, store_type, connector_ref, repo_name)
        if error:
            pipeline_errors.append({
                'org_identifier': org_identifier,
                'project_identifier': project_identifier,
                'pipeline_identifier': pipeline_identifier,
                'error': error
            })
            logging.error(f'Error fetching pipeline YAML: {error} for pipeline {pipeline_identifier}')
            continue

        if pipeline_yaml:
            if pipeline_identifier == DEBUG_PIPELINE_NAME and DEBUG == True:
                logging.info(f'Pipeline YAML for {pipeline_identifier}: {json.dumps(pipeline_yaml, indent=2)}')
            current_level = 'project' if project_identifier else 'org' if org_identifier else 'account'
            if 'template' in pipeline_yaml.get('pipeline', {}) and 'templateRef' in pipeline_yaml['pipeline']['template']:
                template_ref = pipeline_yaml['pipeline']['template']['templateRef']
                if template_ref not in processed_templates:
                    processed_templates[template_ref] = {'count': 0, 'type': None, 'ci': False, 'infra': set()}
                    template_yaml, error = get_template_yaml(template_ref, current_level=current_level, org_identifier=org_identifier, project_identifier=project_identifier, parent_pipeline_id=pipeline_identifier)
                    if template_yaml:
                        if pipeline_identifier == DEBUG_PIPELINE_NAME and DEBUG == True:
                            logging.info(f'Template YAML for {template_ref}: {json.dumps(template_yaml, indent=2)}')
                        processed_templates[template_ref]['count'] += 1
                        logging.info(f'Incremented template count for {template_ref}: {processed_templates[template_ref]["count"]}')
                        template_level = 'account' if template_ref.startswith('account.') else 'org' if template_ref.startswith('org.') else 'project'
                        infra_types_pipeline, ci_stages_count, has_template_stage, templates_used_recursive = process_stages(
                            template_yaml.get('template', {}).get('spec', {}).get('stages', []), processed_templates, template_count, template_level, org_identifier, project_identifier, parent_pipeline_id=pipeline_identifier
                        )
                        infra_types_pipeline = handle_infra_types(infra_types_pipeline)
                        for infra_type in infra_types_pipeline:
                            infra_types[infra_type] += 1
                        total_ci_stages += ci_stages_count
                        if ci_stages_count > 0:
                            ci_stage_count += 1
                            total_pipelines_with_ci += 1
                        if pipeline_identifier == DEBUG_PIPELINE_NAME and DEBUG == True:
                            logging.info(f'Pipeline details after template processing: {pipeline_details}')
                            logging.info(f'Infra types after template processing: {infra_types}')
                        pipeline_details.append({
                            'pipeline_identifier': pipeline_identifier,
                            'org_identifier': org_identifier,
                            'project_identifier': project_identifier,
                            'ci_stages_count': ci_stages_count,
                            'infra_types': ', '.join(infra_types_pipeline),
                            'total_stages': len(pipeline_yaml.get('pipeline', {}).get('stages', [])),
                            'template_count': sum(template_count.values()),
                            'pipeline_name': pipeline.get('name', ''),
                            'templates_used': ', '.join(templates_used_recursive)
                        })
                        continue
                else:
                    # Utilize as informações do template processado anteriormente
                    processed_templates[template_ref]['count'] += 1
                    infra_types_pipeline = processed_templates[template_ref]['infra']
                    ci_stages_count = 1 if processed_templates[template_ref]['ci'] else 0

                    for infra_type in infra_types_pipeline:
                        infra_types[infra_type] += 1
                    total_ci_stages += ci_stages_count
                    if ci_stages_count > 0:
                        ci_stage_count += 1
                        total_pipelines_with_ci += 1

            templates_used_recursive = set()
            infra_types_pipeline, ci_stages_count, has_template_stage, templates_used_recursive = process_stages(
                pipeline_yaml.get('pipeline', {}).get('stages', []), processed_templates, template_count, current_level, org_identifier, project_identifier, parent_pipeline_id=pipeline_identifier
            )
            infra_types_pipeline = handle_infra_types(infra_types_pipeline)
            for infra_type in infra_types_pipeline:
                infra_types[infra_type] += 1
            total_ci_stages += ci_stages_count
            if ci_stages_count > 0:
                ci_stage_count += 1
                total_pipelines_with_ci += 1
            if pipeline_identifier == DEBUG_PIPELINE_NAME and DEBUG == True:
                logging.info(f'Pipeline details after stage processing: {pipeline_details}')
                logging.info(f'Infra types after stage processing: {infra_types}')
            pipeline_details.append({
                'pipeline_identifier': pipeline_identifier,
                'org_identifier': org_identifier,
                'project_identifier': project_identifier,
                'ci_stages_count': ci_stages_count,
                'infra_types': ', '.join(infra_types_pipeline),
                'total_stages': len(pipeline_yaml.get('pipeline', {}).get('stages', [])),
                'template_count': sum(template_count.values()),
                'pipeline_name': pipeline.get('name', ''),
                'templates_used': ', '.join(templates_used_recursive)
            })

    return total_pipelines, total_pipelines_with_ci, total_ci_stages, infra_types, template_count, pipeline_details, pipeline_errors

def handle_infra_types(infra_types_pipeline):
    if len(infra_types_pipeline) > 1:
        infra_types_pipeline = {'Mixed'}
    return infra_types_pipeline


def calculate_percentage(infra_types, total_count):
    if total_count == 0:
        return {k: '0.00%' for k in infra_types}
    
    # Adjust percentages to sum up to 100%
    total_count_adjusted = sum(infra_types.values())
    if total_count_adjusted == 0:
        return {k: '0.00%' for k in infra_types}

    percentages = {k: f'{(v / total_count_adjusted) * 100:.2f}%' for k, v in infra_types.items()}

    return percentages


# Calculate Build time avg and max of pipelines

def fetch_pipeline_executions(org_identifier, project_identifier, pipeline_identifier):
    api_url = 'https://app.harness.io/pipeline/api/pipelines/execution/summary'
    print(f'Fetching pipeline executions for pipeline {pipeline_identifier}: {api_url}')
    payload = {
        "accountIdentifier": "6_vVHzo9Qeu9fXvj-AcbCQ",
        "orgIdentifier": org_identifier,
        "projectIdentifier": project_identifier,
        "pipelineIdentifier": pipeline_identifier,
        "page": 0,
        "size": 20,
        "showAllExecutions": True,
        "module": "CI",
        "getDefaultFromOtherRepo": True,
        "filterType": "PipelineExecution"
    }
    response = requests.post(api_url, headers=headers, data=json.dumps(payload))
    response.raise_for_status()  # Raise an exception for HTTP errors
    return response.json()

def calculate_build_times(executions):
    if not executions:
        return 0, 0
    
    total_time = []
    
    for execution in executions:
        build_time = 0
        for node in execution.get('layoutNodeMap', {}).values():
            if node['nodeType'] == 'CI' and 'startTs' in node and 'endTs' in node:
                build_time += node['endTs'] - node['startTs']
        
        if build_time > 0:
            total_time.append(build_time)
    
    if not total_time:
        return 0, 0
    
    avg_time = sum(total_time) / len(total_time)
    max_time = max(total_time)
    
    return avg_time, max_time

def get_avg_and_max_build_time(org_identifier, project_identifier, pipeline_identifier):
    
    response_data = fetch_pipeline_executions(org_identifier, project_identifier, pipeline_identifier)
    executions = response_data.get('data', {}).get('content', [])
    
    avg_time, max_time = calculate_build_times(executions)
    
    return avg_time, max_time

#  End Calculate build time


def export_to_csv(org_summary, account_summary):
    # Dynamically determine the union of all keys in org_summary to include in the CSV
    all_keys = set()
    for summary in org_summary.values():
        all_keys.update(summary['infra_percentage'].keys())
    
    # Ensure 'VM' is included in the keys if it exists
    all_keys.update(account_summary['infra_percentage'].keys())
    
    fieldnames_account = [
        'total_orgs', 'total_projects', 'total_pipelines', 'total_pipelines_with_ci', 
        'total_ci_stages', 'template_count'
    ] + list(all_keys)

    with open('account_summary.csv', 'w', newline='') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames_account)
        
        writer.writeheader()
        account_summary_data = {
            'total_orgs': account_summary['total_orgs'],
            'total_projects': account_summary['total_projects'],
            'total_pipelines': account_summary['total_pipelines'],
            'total_pipelines_with_ci': account_summary['total_pipelines_with_ci'],
            'total_ci_stages': account_summary['total_ci_stages'],
            'template_count': account_summary['template_count']
        }
        account_summary_data.update(account_summary['infra_percentage'])
        writer.writerow(account_summary_data)

    fieldnames_org = [
        'org_identifier', 'total_pipelines', 'total_pipelines_with_ci', 
        'total_ci_stages', 'template_count'
    ] + list(all_keys)

    with open('org_summary.csv', 'w', newline='') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames_org)

        writer.writeheader()
        for org_identifier, summary in org_summary.items():
            org_summary_data = {
                'org_identifier': org_identifier,
                'total_pipelines': summary['total_pipelines'],
                'total_pipelines_with_ci': summary['total_pipelines_with_ci'],
                'total_ci_stages': summary['total_ci_stages'],
                'template_count': summary['template_count']
            }
            org_summary_data.update(summary['infra_percentage'])
            writer.writerow(org_summary_data)

def export_pipeline_details_to_csv(pipeline_details):
    with open('pipeline_details.csv', 'w', newline='') as csvfile:
        fieldnames = [
            'pipeline_identifier', 'org_identifier', 'project_identifier', 'pipeline_name', 
            'ci_stages_count', 'total_stages', 'template_count', 'templates_used', 'infra_types'
        ]
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

        writer.writeheader()
        for detail in pipeline_details:
            writer.writerow(detail)
            
            # detail_data = {
            #     'org_identifier': detail['org_identifier'],
            #     'project_identifier': detail['project_identifier'],
            #     'pipeline_identifier': detail['pipeline_identifier'],
            #     'pipeline_name': detail['pipeline_name'],
            #     'ci_stages_count': detail['ci_stages_count'],
            #     'total_stages': detail['total_stages'],
            #     'template_count': detail['template_count'],
            #     'templates_used': detail.get('templates_used', '')
            # }
            # detail_data.update(detail['infra_types'])
            # writer.writerow(detail_data)

def export_pipeline_errors_to_csv(pipeline_errors):
    with open('pipeline_errors.csv', 'w', newline='') as csvfile:
        fieldnames = ['org_identifier', 'project_identifier', 'pipeline_identifier', 'error']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

        writer.writeheader()
        for error in pipeline_errors:
            writer.writerow(error)

def export_template_details_to_csv(template_count):
    with open('template_details.csv', 'w', newline='') as csvfile:
        fieldnames = ['template_ref', 'count']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

        writer.writeheader()
        for template_ref, count in template_count.items():
            writer.writerow({'template_ref': template_ref, 'count': count})

def update_spreadsheet(org_summary, account_summary, pipeline_details, template_count_dict):
    file_path = '/Users/diegopereira/Documents/Development/git/serenity/CI-AdoptionPlan-Hosted_Builds_Migration.xlsx'

    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        print("Spreadsheet not found. Creating a new one.")
        workbook = openpyxl.Workbook()

    # Remove the default sheet if it exists
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']

    # Create or update the 'Account Summary' sheet
    if 'Account Summary' not in workbook.sheetnames:
        workbook.create_sheet('Account Summary')
    account_sheet = workbook['Account Summary']

    # Convert account_summary to DataFrame for easier manipulation
    account_summary_df = pd.DataFrame([account_summary])

    # Write the account_summary DataFrame to the sheet
    for r_idx, row in enumerate(dataframe_to_rows(account_summary_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            if isinstance(value, dict):
                value = json.dumps(value)  # Convert dict to JSON string
            account_sheet.cell(row=r_idx, column=c_idx, value=value)

    # Create or update the 'Org Summary' sheet
    if 'Org Summary' not in workbook.sheetnames:
        workbook.create_sheet('Org Summary')
    org_sheet = workbook['Org Summary']

    # Convert org_summary to DataFrame for easier manipulation
    org_summary_df = pd.DataFrame.from_dict(org_summary, orient='index')

    # Write the org_summary DataFrame to the sheet
    for r_idx, row in enumerate(dataframe_to_rows(org_summary_df, index=True, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            if isinstance(value, dict):
                value = json.dumps(value)  # Convert dict to JSON string
            org_sheet.cell(row=r_idx, column=c_idx, value=value)

    # Create or update the 'Pipeline Details' sheet
    if 'Pipeline Details' not in workbook.sheetnames:
        workbook.create_sheet('Pipeline Details')
    pipeline_sheet = workbook['Pipeline Details']

    # Convert pipeline_details to DataFrame for easier manipulation
    pipeline_details_df = pd.DataFrame(pipeline_details)

    # Write the pipeline_details DataFrame to the sheet
    for r_idx, row in enumerate(dataframe_to_rows(pipeline_details_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            pipeline_sheet.cell(row=r_idx, column=c_idx, value=value)

    # Create or update the 'Template Details' sheet
    if 'Template Details' not in workbook.sheetnames:
        workbook.create_sheet('Template Details')
    template_sheet = workbook['Template Details']

    # Ensure template_count_dict is a dictionary
    if not isinstance(template_count_dict, dict):
        raise ValueError("template_count_dict should be a dictionary")

    # Convert template_count_dict to DataFrame for easier manipulation
    template_details_df = pd.DataFrame(list(template_count_dict.items()), columns=['template_ref', 'count'])

    # Write the template_details DataFrame to the sheet
    for r_idx, row in enumerate(dataframe_to_rows(template_details_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            template_sheet.cell(row=r_idx, column=c_idx, value=value)

    # Save the updated workbook
    new_file_path = '/Users/diegopereira/Documents/Development/git/serenity/CI-AdoptionPlan-Hosted_Builds_Migration_Updated.xlsx'
    workbook.save(new_file_path)
    print(f"Spreadsheet updated and saved to {new_file_path}")




def main():
    orgs = get_orgs()
    total_orgs = len(orgs)
    total_projects = 0
    total_pipelines = 0
    total_pipelines_with_ci = 0
    total_ci_stages = 0
    template_count = defaultdict(int)
    infra_types_account = defaultdict(int)
    org_summary = {}
    pipeline_details = []
    pipeline_errors = []
    processed_templates = {}
    template_count_dict = defaultdict(int)

    for org in orgs:
        org_identifier = org['organization']['identifier']
        print(f'Processing org: {org_identifier}')
        projects = get_projects(org_identifier)
        total_projects += len(projects)

        infra_types_org = defaultdict(int)
        ci_stage_count_org = 0
        total_ci_stages_org = 0
        template_count_org = defaultdict(int)
        total_pipelines_org = 0

        for project in projects:
            project_identifier = project['projectResponse']['project']['identifier']
            org_identifier = project['projectResponse']['project']['orgIdentifier']
            print(f'Processing project: {project_identifier} in org: {org_identifier}')
            pipelines = get_pipelines(org_identifier, project_identifier)
            if pipelines:
                total_pipelines_org += len(pipelines)
                total_pipelines_project, ci_pipelines_count, total_stages, infra_types, template_count_local, details, errors = analyze_pipelines(
                    pipelines, org_identifier, project_identifier, processed_templates, template_count_dict
                )
                total_pipelines += total_pipelines_project
                total_pipelines_with_ci += ci_pipelines_count
                total_ci_stages += total_stages
                for template_ref, count in template_count_local.items():
                    template_count[template_ref] += count
                ci_stage_count_org += ci_pipelines_count
                total_ci_stages_org += total_stages
                for template_ref, count in template_count_local.items():
                    template_count_org[template_ref] += count
                for infra_type, count in infra_types.items():
                    infra_types_org[infra_type] += count
                pipeline_details.extend(details)
                pipeline_errors.extend(errors)

        infra_percentage_org = calculate_percentage(infra_types_org, ci_stage_count_org)
        org_summary[org_identifier] = {
            'total_pipelines': total_pipelines_org,
            'total_pipelines_with_ci': ci_stage_count_org,
            'total_ci_stages': total_ci_stages_org,
            'template_count': dict(template_count_org),
            'infra_percentage': infra_percentage_org
        }

        for infra_type, count in infra_types_org.items():
            infra_types_account[infra_type] += count

    infra_percentage_account = calculate_percentage(infra_types_account, total_pipelines_with_ci)
    avg_pipelines_per_project = total_pipelines / total_projects if total_projects > 0 else 0
    avg_projects_per_org = total_projects / total_orgs if total_orgs > 0 else 0

    account_summary = {
        'total_orgs': total_orgs,
        'total_projects': total_projects,
        'total_pipelines': total_pipelines,
        'total_pipelines_with_ci': total_pipelines_with_ci,
        'total_ci_stages': total_ci_stages,
        'template_count': dict(template_count),
        'infra_percentage': infra_percentage_account
    }

    print(f'Total Organizations: {total_orgs}')
    print(f'Total Projects: {total_projects}')
    print(f'Total Pipelines: {total_pipelines}')
    print(f'Pipelines with CI Stage: {total_pipelines_with_ci}')
    print(f'Total CI Stages: {total_ci_stages}')
    print(f'Templates in Pipelines: {sum(template_count.values())}')
    print(f'Average Pipelines per Project: {avg_pipelines_per_project:.2f}')
    print(f'Average Projects per Organization: {avg_projects_per_org:.2f}')
    print(f'\nInfrastructure types for account:')
    for infra_type, percentage in infra_percentage_account.items():
        print(f'{infra_type}: {percentage}')

    print('\nInfrastructure types by org:')
    for org_identifier, summary in org_summary.items():
        print(f'\nOrg: {org_identifier}')
        print(f'Total Pipelines: {summary["total_pipelines"]}')
        print(f'Pipelines with CI Stage: {summary["total_pipelines_with_ci"]}')
        print(f'CI Stage Count: {summary["total_ci_stages"]}')
        print(f'Templates in Pipelines: {sum(summary["template_count"].values())}')
        for infra_type, percentage in summary['infra_percentage'].items():
            print(f'{infra_type}: {percentage}')

    export_to_csv(org_summary, account_summary)
    export_pipeline_details_to_csv(pipeline_details)
    export_pipeline_errors_to_csv(pipeline_errors)
    export_template_details_to_csv(template_count_dict)
    update_spreadsheet(org_summary, account_summary, pipeline_details, template_count_dict)

if __name__ == "__main__":
    main()
